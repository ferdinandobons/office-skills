# SPDX-License-Identifier: MIT
"""Deterministic builder for the COMPLEX synthetic XLSX example template.

Produces ``examples/templates/branddocs_template.xlsx``: a 100% synthetic
(BrandDocs-style, never proprietary) workbook that stresses the brand-xlsx engine
across as many Excel component types as openpyxl can author:

  * MULTIPLE sheets in a deliberate tab order (Cover, Inputs, Model, Summary,
    Dashboard, Scenarios, Data) - exercises multi-sheet structure + skeleton ordering.
  * NAMED regions of every geometry the extractor distinguishes:
      - a single-cell title anchor sitting under a MERGED header row,
      - a single-cell that sits in a FROZEN header band,
      - a multi-cell INPUTS block (sample-data candidate),
      - a multi-cell DATA region that is the BODY of a native table object,
      - a single named cell that is a cross-sheet formula target.
  * FORMULAS: in-sheet SUM/total rows, percent-of-total ratios, a cross-sheet
    reference (Summary pulls from Model and Inputs), and a SUBTOTAL - all
    authored verbatim so we can later assert they survive generation byte-exact.
  * NUMBER FORMATS: currency (accounting), percent, thousands, and ISO date.
  * A native TABLE object (``BrandDocsDataTbl``) with a banded table style.
  * A dashboard sheet with KPI cards, cross-sheet formulas, and a third native
    chart so the rendered workbook looks like a finished executive model.
  * A scenarios sheet with list validation and a native scenario table.
  * CONDITIONAL FORMATTING: a color scale, a cell-is rule, and a formula rule.
  * FROZEN PANES on the data sheets.
  * The shared text-only BrandDocs wordmark placed as a worksheet drawing on the
    cover - rendered in-process by ``_brandlib.branddocs_mark_png`` (no
    external/proprietary asset on disk), sized at 4:1 so the text is not
    distorted.
  * Named cell STYLES (``BrandDocsTitle``, ``BrandDocsHeader``, ``BrandDocsCurrency``,
    ``BrandDocsPercent``, ``BrandDocsInput``) registered on the workbook and applied.
  * Demo / sample data rows the generator is expected to clear.

The output is byte-reproducible: a fixed timestamp / fixed image bytes / sorted
defined names, so re-running the builder yields an identical file (CI-friendly).

Run:
    PYTHONPATH=scripts .venv/bin/python examples/builders/build_branddocs_xlsx.py
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
    Rule,
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableStyleInfo

from _brandlib import branddocs_mark_png, brand_theme_slots, freeze_ooxml
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).resolve().parents[1] / "templates" / "branddocs_template.xlsx"

# Synthetic BrandDocs brand palette (made-up; never proprietary).
BRAND_NAVY = "FF16213F"
BRAND_TEAL = "FF2B7CD3"
BRAND_AMBER = "FFE0742B"
BRAND_LIGHT = "FFEAF1FF"
WHITE = "FFFFFFFF"

# 6-digit (no alpha) variants for openpyxl chart series fills, derived from the
# brand palette constants above (no new brand literals introduced).
NAVY6 = BRAND_NAVY[-6:]  # "16213F"
TEAL6 = BRAND_TEAL[-6:]  # "2B7CD3"
AMBER6 = BRAND_AMBER[-6:]  # "E0742B"


# ---------------------------------------------------------------------------
# Named cell styles (registered once on the workbook).
# ---------------------------------------------------------------------------
def _register_named_styles(wb: Workbook) -> None:
    thin = Side(style="thin", color=BRAND_NAVY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = NamedStyle(name="BrandDocsTitle")
    title.font = Font(name="Arial", size=20, bold=True, color=BRAND_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")

    header = NamedStyle(name="BrandDocsHeader")
    header.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    header.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.border = border

    currency = NamedStyle(name="BrandDocsCurrency")
    currency.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    currency.font = Font(name="Calibri", size=11, color=BRAND_NAVY)
    currency.border = border

    percent = NamedStyle(name="BrandDocsPercent")
    percent.number_format = "0.0%"
    percent.font = Font(name="Calibri", size=11, color=BRAND_TEAL)
    percent.border = border

    inp = NamedStyle(name="BrandDocsInput")
    inp.fill = PatternFill("solid", fgColor=BRAND_LIGHT)
    inp.font = Font(name="Calibri", size=11, color=BRAND_NAVY)
    inp.border = border

    kpi = NamedStyle(name="BrandDocsKPI")
    kpi.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    kpi.font = Font(name="Arial", size=15, bold=True, color=WHITE)
    kpi.alignment = Alignment(horizontal="center", vertical="center")
    kpi.border = border

    # Additional reusable brand cell styles (a complete date / multiple / total /
    # subtitle / band-header system). Each is applied to a real cell below so the
    # extractor promotes it into a cell_style role with a font/fill/border digest.
    top_navy = Border(top=Side(style="thin", color=BRAND_NAVY))

    subtitle = NamedStyle(name="BrandDocsSubtitle")
    subtitle.font = Font(name="Arial", size=12, italic=True, color=BRAND_TEAL)
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    total = NamedStyle(name="BrandDocsTotal")
    total.font = Font(name="Arial", size=11, bold=True, color=BRAND_NAVY)
    total.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    total.border = top_navy

    date = NamedStyle(name="BrandDocsDate")
    date.font = Font(name="Calibri", size=11, color=BRAND_NAVY)
    date.number_format = "yyyy-mm-dd"

    multiple = NamedStyle(name="BrandDocsMultiple")
    multiple.font = Font(name="Calibri", size=11, color=BRAND_NAVY)
    multiple.number_format = "0.00x"

    band_header = NamedStyle(name="BrandDocsBandHeader")
    band_header.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    band_header.fill = PatternFill("solid", fgColor=BRAND_NAVY)
    band_header.alignment = Alignment(horizontal="left", vertical="center")

    for style in (
        title,
        header,
        currency,
        percent,
        inp,
        kpi,
        subtitle,
        total,
        date,
        multiple,
        band_header,
    ):
        wb.add_named_style(style)


# ---------------------------------------------------------------------------
# Per-sheet brand chrome (slim navy title band + teal rule + print header/footer).
# ---------------------------------------------------------------------------
def _brand_band(ws, title: str, *, last_col: int = 7, band_row: int = 1) -> None:
    """Stamp a slim navy brand band (white sheet-title) + a teal rule below it.

    Pure static fills/merges on a fixed row pair, so it surfaces in the
    ``non_empty_cells`` / ``merged_cells`` inventories the package walker reads.
    The band sits ABOVE the sheet's existing content rows (the body builders below
    start at row 3+), so it is additive and never overwrites a formula/header.
    """
    end = get_column_letter(last_col)
    ws.merge_cells(f"A{band_row}:{end}{band_row}")
    title_cell = ws.cell(row=band_row, column=1, value=title)
    title_cell.style = "BrandDocsBandHeader"
    for col in range(1, last_col + 1):
        ws.cell(row=band_row, column=col).fill = PatternFill(
            "solid", fgColor=BRAND_NAVY
        )
    ws.row_dimensions[band_row].height = 22
    rule_row = band_row + 1
    for col in range(1, last_col + 1):
        ws.cell(row=rule_row, column=col).fill = PatternFill(
            "solid", fgColor=BRAND_TEAL
        )
    ws.row_dimensions[rule_row].height = 4


def _print_chrome(ws, *, title_rows: str = "1:3") -> None:
    """Set print header/footer live fields + repeating print-title rows.

    ``&A`` (sheet name) centred, ``&D`` (date) left, ``Page &P of &N`` right -
    literal Excel field codes (no wall-clock), and ``print_title_rows`` so the
    brand band repeats on every printed page. Structural print metadata only.
    """
    ws.oddHeader.center.text = "&A"
    ws.oddHeader.left.text = "&D"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.print_title_rows = title_rows


# ---------------------------------------------------------------------------
# Sheet builders.
# ---------------------------------------------------------------------------
def _build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    # Merged title band A1:G1 with a single-cell named title anchor at A1.
    ws.merge_cells("A1:G1")
    # Title band: navy fill + white title text so the cover reads as branded.
    ws["A1"] = "FY2025 Revenue Performance Review"
    ws["A1"].style = "BrandDocsTitle"
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=BRAND_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:G2")
    ws["A2"] = "Quarterly revenue model and executive summary"
    ws["A2"].style = "BrandDocsSubtitle"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A4"] = "Prepared for"
    ws["B4"] = "BrandDocs Corp (synthetic demo)"
    ws["A5"] = "Reporting period"
    ws["B5"] = "FY2025 (Q1-Q4)"
    # An ISO-date cell carrying the reusable BrandDocsDate brand style (navy
    # Calibri + yyyy-mm-dd mask), pinned and deterministic.
    ws["A6"] = "Generated on"
    ws["B6"] = "2026-01-15"
    ws["B6"].style = "BrandDocsDate"
    # Three scorecard tiles make the cover read as an executive-ready workbook.
    scorecards = [
        ("D4", "FY net revenue", "=Summary!B4"),
        ("E4", "Net margin", "=Summary!B8"),
        ("F4", "Brand audit", "Deep QA"),
    ]
    for anchor, label, value in scorecards:
        col = ws[anchor].column
        ws.cell(row=4, column=col, value=label).style = "BrandDocsHeader"
        vc = ws.cell(row=5, column=col, value=value)
        vc.style = "BrandDocsKPI"
        if value.startswith("="):
            vc.number_format = "0.0%" if "B8" in value else "#,##0"
    ws["G4"] = "Scope"
    ws["G4"].style = "BrandDocsHeader"
    ws["G5"] = "3 formats"
    ws["G5"].style = "BrandDocsKPI"
    ws["G5"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    # A navy brand band across the printable width, below the scorecards.
    ws.merge_cells("A8:G8")
    for col in range(1, 8):
        ws.cell(row=8, column=col).fill = PatternFill("solid", fgColor=BRAND_NAVY)
    ws.row_dimensions[8].height = 28
    # Header drawing: the shared BrandDocs wordmark sits inside the printable
    # area at a 4:1 aspect ratio so the text is not distorted.
    logo = XLImage(_logo_path())
    logo.width, logo.height = 180, 45
    ws.add_image(logo, "A9")
    # A small color-code LEGEND describing the model conventions (real financial
    # models annotate their color code). Styled cells, so it surfaces as
    # non-empty cells the package walker reads.
    ws["D11"] = "Legend"
    ws["D11"].style = "BrandDocsBandHeader"
    ws.merge_cells("D11:G11")
    ws["D12"] = "Inputs"
    ws["D12"].style = "BrandDocsInput"
    ws["E12"] = "Blue fill = editable assumptions you fill in"
    ws["E12"].font = Font(name="Calibri", size=10, color=BRAND_NAVY)
    ws["D13"] = "Formulas"
    ws["D13"].font = Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY)
    ws["E13"] = "Navy text = model formulas (do not overwrite)"
    ws["E13"].font = Font(name="Calibri", size=10, color=BRAND_NAVY)
    # Document the headline scorecard with a source note.
    ws["D5"].comment = Comment(
        "Synthetic figure - pulled live from the Summary tab.", "BrandDocs"
    )
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    for col in "CDEFG":
        ws.column_dimensions[col].width = 16
    for row in (4, 5):
        ws.row_dimensions[row].height = 28
    # Keep the whole cover on a single print page.
    ws.print_area = "A1:G13"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _build_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "Model Inputs"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=BRAND_NAVY)
    # Header row 3.
    for col, label in enumerate(("Driver", "Value", "Unit"), start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.style = "BrandDocsHeader"
    # Input block rows 4..7 (the named INPUTS region, demo values).
    # Reconciled so Implied gross (Units * Price) == Model Gross revenue FY
    # (29070 * 50 = 1,453,500), giving Net margin = 85.0% (sensible range).
    inputs = [
        ("Units sold", 29070, "ea"),
        ("Unit price", 50.0, "USD"),
        ("Discount rate", 0.12, "pct"),
        ("Tax rate", 0.22, "pct"),
    ]
    for i, (driver, val, unit) in enumerate(inputs):
        r = 4 + i
        ws.cell(row=r, column=1, value=driver).style = "BrandDocsInput"
        vc = ws.cell(row=r, column=2, value=val)
        vc.style = "BrandDocsInput"
        if unit == "USD":
            vc.style = "BrandDocsCurrency"
        elif unit == "pct":
            vc.style = "BrandDocsPercent"
        elif unit == "ea":
            # Whole-unit counts get a thousands separator so "Units sold" reads
            # "29,070" instead of "29070" (keeps the input fill/border style).
            vc.number_format = "#,##0"
        ws.cell(row=r, column=3, value=unit).style = "BrandDocsInput"
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    # Distributed-template convention: lock the sheet but leave ONLY the input
    # value cells (B4:B7) editable, so a recipient fills the blue inputs and cannot
    # disturb the model. (openpyxl honors locked/unlocked on save; enforcement is
    # Excel-side, so the generator can still fill the named region programmatically.)
    for r in range(4, 8):
        ws.cell(row=r, column=2).protection = Protection(locked=False)
    ws.protection.sheet = True
    # Guided-input validations with prompts + error alerts (a fillable model).
    # Units sold: whole number >= 0.
    dv_units = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True,
        promptTitle="Units sold",
        prompt="Enter the synthetic unit volume (whole number, 0 or more).",
        errorTitle="Invalid units",
        error="Units sold must be a whole number greater than or equal to 0.",
    )
    ws.add_data_validation(dv_units)
    dv_units.add(ws["B4"])
    # Discount + Tax rate: decimal between 0 and 1.
    dv_rate = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True,
        promptTitle="Rate (0-1)",
        prompt="Enter a rate as a fraction between 0 and 1 (e.g. 0.12 for 12%).",
        errorTitle="Invalid rate",
        error="Rate must be a decimal between 0 and 1.",
    )
    ws.add_data_validation(dv_rate)
    dv_rate.add(ws["B6"])
    dv_rate.add(ws["B7"])
    # CF (blanks): flag any empty input value cell so a recipient sees what is
    # still missing (a brand-light fill on B4:B7 when blank).
    ws.conditional_formatting.add(
        "B4:B7",
        Rule(
            type="containsBlanks",
            formula=["LEN(TRIM(B4))=0"],
            dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=BRAND_LIGHT)),
        ),
    )
    # Print-ready header/footer with live fields (sheet name + page x of y).
    ws.oddHeader.center.text = "&A"
    ws.oddFooter.right.text = "Page &P of &N"
    # Source notes on every input value cell (self-documenting model).
    for r in range(4, 8):
        ws.cell(row=r, column=2).comment = Comment(
            "Synthetic input - replace with your data.", "BrandDocs"
        )


def _build_model(wb: Workbook) -> None:
    ws = wb.create_sheet("Model")
    ws["A1"] = "Revenue Model"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=BRAND_NAVY)
    # Header row 3 for the native table.
    headers = ("Line item", "Q1", "Q2", "Q3", "Q4", "FY Total", "% of FY")
    for col, label in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=label).style = "BrandDocsHeader"
    # Demo body rows 4..7 with cross-quarter SUM + percent-of-total formulas.
    body = [
        ("Gross revenue", 320000, 351000, 372500, 410000),
        ("Discounts", -38400, -42120, -44700, -49200),
        ("Returns", -9600, -10530, -11175, -12300),
        ("Net revenue", None, None, None, None),  # formula row below
    ]
    for i, row in enumerate(body):
        r = 4 + i
        ws.cell(row=r, column=1, value=row[0])
        if row[0] == "Net revenue":
            # Net revenue = sum of the three lines above, per quarter.
            for col in range(2, 6):
                L = get_column_letter(col)
                c = ws.cell(row=r, column=col, value=f"=SUM({L}4:{L}6)")
                c.number_format = "#,##0"
        else:
            for col, val in zip(range(2, 6), row[1:]):
                c = ws.cell(row=r, column=col, value=val)
                c.number_format = "#,##0"
        # FY Total (col 6) = SUM of the four quarters in this row.
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})").number_format = "#,##0"
        # % of FY (col 7) = this row's FY total / net-revenue FY total (row 7).
        ws.cell(
            row=r, column=7, value=f"=IF($F$7=0,0,F{r}/$F$7)"
        ).number_format = "0.0%"
    # Table TOTALS row 8: a SUBTOTAL over the FY-total column, carried inside the
    # table object (showTotalsRow) so the native table gains a labelled total row.
    ws.cell(row=8, column=1, value="Total").style = "BrandDocsTotal"
    ws.cell(row=8, column=6, value="=SUBTOTAL(109,F4:F7)").style = "BrandDocsTotal"
    # A grand-total SUBTOTAL row 10 (col 6) carrying the brand total style.
    ws.cell(row=10, column=1, value="Subtotal (visible)").style = "BrandDocsTotal"
    sub = ws.cell(row=10, column=6, value="=SUBTOTAL(9,F4:F6)")
    sub.style = "BrandDocsTotal"
    sub.number_format = "#,##0"
    # A STRUCTURED-REFERENCE formula that reads the table's FY Total column.
    ws.cell(
        row=10, column=7, value="=SUM(BrandDocsDataTbl[FY Total])"
    ).number_format = "#,##0"
    # Native TABLE object over the body (header row 3 .. data row 7 .. totals row 8).
    table = Table(displayName="BrandDocsDataTbl", ref="A3:G8")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    # Mark the table's totals row so Excel renders row 8 as the table total band.
    table.totalsRowShown = True
    table.totalsRowCount = 1
    # The autofilter must span the header + data rows ONLY (A3:G7); the totals row
    # (row 8) is excluded, else Excel reports a repair-needed table.
    table.autoFilter = AutoFilter(ref="A3:G7")
    ws.add_table(table)
    # The totals-row label + the FY-total SUBTOTAL are authored as real cells in
    # row 8 above; the table's totalsRow band (totalsRowShown/Count) renders them.
    # CONDITIONAL FORMATTING: color scale on the quarter grid, a CellIs rule on
    # the % column, and a formula rule that flags negative FY totals.
    ws.conditional_formatting.add(
        "B4:E6",
        ColorScaleRule(
            start_type="min",
            start_color="FFF8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFEB84",
            end_type="max",
            end_color="FF63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        "G4:G7",
        CellIsRule(
            operator="greaterThan",
            formula=["0.5"],
            fill=PatternFill("solid", fgColor=BRAND_AMBER),
        ),
    )
    ws.conditional_formatting.add(
        "F4:F6",
        FormulaRule(formula=["F4<0"], fill=PatternFill("solid", fgColor="FFFFC7CE")),
    )
    # Two more common CF families (icon set + top-N), so the CF inventory the model
    # reads covers iconSet/top10 too, not just colorScale/cellIs/expression/dataBar.
    ws.conditional_formatting.add(
        "G4:G6",
        IconSetRule(icon_style="3TrafficLights1", type="percent", values=[0, 33, 67]),
    )
    ws.conditional_formatting.add(
        "B4:E6",
        Rule(
            type="top10",
            rank=3,
            dxf=DifferentialStyle(fill=PatternFill("solid", fgColor="FF63BE7B")),
        ),
    )
    # Negative line items (Discounts/Returns) use a red-negative accounting mask -
    # a distinct surfaced number format, on existing cells (no geometry change).
    for r in (5, 6):
        for col in range(2, 7):
            ws.cell(row=r, column=col).number_format = "#,##0;[Red](#,##0)"
    ws.freeze_panes = "B4"
    # Wider label column, tighter numeric columns so the table + chart fit a page.
    ws.column_dimensions["A"].width = 16
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # A native CHART driven by the model Net revenue row (one series, 4 points
    # across the Q1..Q4 categories).
    chart = BarChart()
    chart.title = "Quarterly net revenue"
    chart.type = "col"
    data = Reference(ws, min_col=2, min_row=7, max_col=5, max_row=7)  # B7:E7
    cats = Reference(ws, min_col=2, min_row=3, max_col=5, max_row=3)  # Q1..Q4
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)
    # Brand: single navy/teal series, no redundant legend, clean axes.
    chart.series[0].graphicalProperties.solidFill = TEAL6
    chart.legend = None
    chart.y_axis.numFmt = "#,##0"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.width = 14
    chart.height = 8
    # Anchor the chart below the table so it stays within the print page.
    ws.add_chart(chart, "A13")
    # Contain the sheet on a single landscape page.
    ws.print_area = "A1:G29"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Repeating print header/footer with live fields on every printed page.
    _print_chrome(ws, title_rows="1:3")

    # A second chart (line) on the Summary sheet is added there.


def _build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Executive Summary"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=BRAND_NAVY)
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["A3"].style = "BrandDocsHeader"
    ws["B3"].style = "BrandDocsHeader"
    # CROSS-SHEET formulas: Summary pulls from Model and Inputs.
    ws["A4"] = "FY net revenue"
    ws["B4"] = "=Model!F7"
    ws["B4"].number_format = "#,##0"
    ws["A5"] = "Units sold (input)"
    ws["B5"] = "=Inputs!B4"
    ws["B5"].number_format = "#,##0"  # thousands separator: "29,070"
    ws["A6"] = "Unit price (input)"
    ws["B6"] = "=Inputs!B5"
    ws["B6"].style = "BrandDocsCurrency"
    ws["A7"] = "Implied gross"
    ws["B7"] = "=Inputs!B4*Inputs!B5"
    ws["B7"].style = "BrandDocsCurrency"
    ws["A8"] = "Net margin"
    ws["B8"] = "=IF(B7=0,0,B4/B7)"
    ws["B8"].style = "BrandDocsPercent"
    ws["A9"] = "Headline KPI"
    ws["B9"] = "85% net margin on $1.45M gross"  # single-cell named output slot
    ws["B9"].comment = Comment(
        "Synthetic headline - derived from the model and inputs tabs.", "BrandDocs"
    )
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    # CF (top10): highlight the single largest metric value among the numeric rows.
    ws.conditional_formatting.add(
        "B4:B8",
        Rule(
            type="top10",
            rank=1,
            dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=BRAND_LIGHT)),
        ),
    )
    # Repeat a small BrandDocs wordmark on the summary so the printed deliverable
    # carries the brand identity beyond the cover (reuses the cached deterministic
    # wordmark PNG; no fresh raster).
    mark = XLImage(_logo_path())
    mark.width, mark.height = 120, 30
    ws.add_image(mark, "D1")

    # A LINE chart on the summary tracking Net revenue per quarter (cross-sheet
    # reference into the Model sheet), coherent with the Model bar chart.
    model = wb["Model"]
    chart = LineChart()
    chart.title = "Net revenue trend (Q1-Q4)"
    data = Reference(model, min_col=2, min_row=7, max_col=5, max_row=7)  # B7:E7
    cats = Reference(model, min_col=2, min_row=3, max_col=5, max_row=3)  # Q1..Q4
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)
    # Name the single series so the legend reads sensibly.
    chart.series[0].tx = SeriesLabel(v="Net revenue")
    # Brand: amber line, clean axes, legend below to avoid extra width.
    chart.series[0].graphicalProperties.line.solidFill = AMBER6
    chart.series[0].graphicalProperties.line.width = 28000  # EMU (~2.2pt)
    chart.series[0].smooth = False
    chart.y_axis.numFmt = "#,##0"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend.position = "b"
    chart.width = 14
    chart.height = 8
    # Anchor below the metric table, keep on a single page.
    ws.add_chart(chart, "A11")
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    _print_chrome(ws, title_rows="1:3")


def _build_dashboard(wb: Workbook) -> None:
    """Executive dashboard with formula-backed KPI cards and a native chart."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Executive Dashboard"
    ws["A1"].font = Font(name="Arial", size=18, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Formula-backed snapshot generated from the model, inputs, and summary tabs"
    )
    ws["A2"].font = Font(name="Arial", size=11, italic=True, color=BRAND_TEAL)
    ws.merge_cells("A2:G2")

    kpis = [
        (1, "FY net revenue", "=Summary!B4", "#,##0"),
        (3, "Net margin", "=Summary!B8", "0.0%"),
        (5, "Input units", "=Summary!B5", "#,##0"),
        (7, "Scenario", "=Scenarios!B3", "@"),
    ]
    for start_col, label, formula, number_format in kpis:
        ws.merge_cells(
            start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1
        )
        ws.merge_cells(
            start_row=5, start_column=start_col, end_row=5, end_column=start_col + 1
        )
        cell = ws.cell(row=4, column=start_col, value=label)
        cell.style = "BrandDocsHeader"
        value_cell = ws.cell(row=5, column=start_col, value=formula)
        value_cell.style = "BrandDocsKPI"
        value_cell.number_format = number_format
        value_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws["A7"] = "Quarter"
    ws["B7"] = "Net revenue"
    ws["C7"] = "Growth"
    for cell in ws["A7:C7"][0]:
        cell.style = "BrandDocsHeader"
    for i, q in enumerate(("Q1", "Q2", "Q3", "Q4"), start=8):
        qidx = i - 7
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=f"=Model!{get_column_letter(qidx + 1)}7")
        ws.cell(row=i, column=2).number_format = "#,##0"
        if i == 8:
            ws.cell(row=i, column=3, value="-")
        else:
            ws.cell(row=i, column=3, value=f"=IF(B{i - 1}=0,0,B{i}/B{i - 1}-1)")
            ws.cell(row=i, column=3).number_format = "0.0%"
    ws.conditional_formatting.add(
        "B8:B11",
        DataBarRule(start_type="min", end_type="max", color=TEAL6, showValue=True),
    )
    # CF (icon set) on the per-quarter growth column so up/flat/down reads at a
    # glance (a second CF family on this sheet alongside the existing data bar).
    ws.conditional_formatting.add(
        "C9:C11",
        IconSetRule(icon_style="3Arrows", type="percent", values=[0, 33, 67]),
    )

    chart = BarChart()
    chart.title = "Net revenue"
    chart.add_data(
        Reference(ws, min_col=2, min_row=7, max_row=11), titles_from_data=True
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=8, max_row=11))
    chart.series[0].graphicalProperties.solidFill = TEAL6
    chart.legend = None
    chart.width = 8.5
    chart.height = 6
    ws.add_chart(chart, "E8")

    # KPI trend mini-table (rows 15..17): one row per KPI with its Q1..Q4 cells,
    # used as the data range for in-cell line sparklines (column F) and richer
    # number-format masks (a signed-percent growth cell and a bps margin-delta).
    ws["A14"] = "KPI trends"
    ws["A14"].style = "BrandDocsBandHeader"
    trend_headers = ("KPI", "Q1", "Q2", "Q3", "Q4", "Trend")
    for col, label in enumerate(trend_headers, start=1):
        ws.cell(row=15, column=col, value=label).style = "BrandDocsHeader"
    trend_rows = [
        ("Net revenue", 272000, 298350, 316625, 348500),
        ("Bookings", 240000, 281000, 305000, 366000),
        ("Active seats", 1180, 1240, 1305, 1402),
    ]
    for r, (label, *quarters) in enumerate(trend_rows, start=16):
        ws.cell(row=r, column=1, value=label)
        for c, q in enumerate(quarters, start=2):
            ws.cell(row=r, column=c, value=q).number_format = "#,##0"
    # Signed-percent growth + a bps margin-delta, exercising two more masks
    # ('+0.0%;-0.0%' and '#,##0 "bps"') on real cells so the family inventory
    # widens beyond currency/percent/thousands/date/multiple. Placed in column G
    # of the trend block (row 16/17), clear of the merged KPI cards above.
    ws["G15"] = "QoQ / delta"
    ws["G15"].style = "BrandDocsHeader"
    ws["G16"] = "=IF(B11=0,0,B11/B8-1)"
    ws["G16"].number_format = "+0.0%;-0.0%"
    ws["G17"] = "=ROUND((Summary!B8-0.85)*10000,0)"
    ws["G17"].number_format = '#,##0 "bps"'

    # Column D carries the KPI-trend Q3 numbers (and its "Q3" header), so it must
    # be wide enough to avoid "###" overflow / clipped headers; keep it consistent
    # with the other data columns at width 12 rather than the old narrow gutter.
    for col, width in zip("ABCDEFGH", (14, 12, 12, 12, 12, 12, 12, 12)):
        ws.column_dimensions[col].width = width
    ws.print_area = "A1:H24"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    _print_chrome(ws, title_rows="1:2")


def _build_scenarios(wb: Workbook) -> None:
    ws = wb.create_sheet("Scenarios")
    ws["A1"] = "Scenario Controls"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=BRAND_NAVY)
    ws["A3"] = "Selected scenario"
    ws["B3"] = "Base"
    ws["B3"].style = "BrandDocsInput"
    dv = DataValidation(
        type="list", formula1='"Base,Upside,Downside"', allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add(ws["B3"])
    ws["B3"].comment = Comment(
        "Synthetic scenario selector used by formulas and visual QA.", "BrandDocs"
    )
    headers = ("Scenario", "Revenue multiplier", "Margin delta", "Narrative")
    for col, label in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=label).style = "BrandDocsHeader"
    rows = [
        ("Base", 1.00, 0.000, "Plan case"),
        ("Upside", 1.12, 0.025, "Faster adoption"),
        ("Downside", 0.91, -0.035, "Cost pressure"),
    ]
    for r, row in enumerate(rows, start=6):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:
                # The revenue multiplier carries the reusable BrandDocsMultiple
                # brand style (navy Calibri + '0.00x' valuation-multiple mask).
                cell.style = "BrandDocsMultiple"
            elif c == 3:
                cell.style = "BrandDocsPercent"
                cell.number_format = "0.0%"
    # A custom-formula validation on the selector keeps it within the listed set
    # (guards a typed value), with an input prompt + error alert.
    dv_custom = DataValidation(
        type="custom",
        formula1="=COUNTIF($A$6:$A$8,$B$3)>0",
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True,
        promptTitle="Scenario",
        prompt="Pick a scenario that exists in the scenario table below.",
        errorTitle="Unknown scenario",
        error="The scenario must match one of Base / Upside / Downside.",
    )
    ws.add_data_validation(dv_custom)
    dv_custom.add(ws["B3"])
    ws["A10"] = "Scenario revenue"
    ws["B10"] = "=Summary!B4*INDEX(B6:B8,MATCH($B$3,A6:A8,0))"
    ws["B10"].number_format = "#,##0"
    ws["A11"] = "Scenario margin"
    ws["B11"] = "=Summary!B8+INDEX(C6:C8,MATCH($B$3,A6:A8,0))"
    ws["B11"].number_format = "0.0%"
    table = Table(displayName="BrandDocsScenarioTbl", ref="A5:D8")
    # The Scenarios table wears the CUSTOM BrandDocsTableStyle (branded dxf header /
    # stripe / total bands); the Model + Dashboard tables stay on TableStyleMedium2
    # so both a custom and a built-in table style appear in the inventory.
    table.tableStyleInfo = TableStyleInfo(
        name="BrandDocsTableStyle",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(table)
    for col, width in zip("ABCD", (18, 18, 16, 24)):
        ws.column_dimensions[col].width = width
    ws.print_area = "A1:D13"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    _print_chrome(ws, title_rows="1:5")


def _build_data(wb: Workbook) -> None:
    """A raw transactions sheet (demo data the generator should clear/refill)."""
    ws = wb.create_sheet("Data")
    headers = ("Date", "Region", "Product", "Amount")
    for col, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=label).style = "BrandDocsHeader"
    demo = [
        ("2025-10-01", "North", "Widget", 12500),
        ("2025-10-02", "South", "Gadget", 9800),
        ("2025-10-03", "East", "Widget", 14200),
    ]
    for i, (d, region, prod, amt) in enumerate(demo):
        r = 2 + i
        # The date column carries the reusable BrandDocsDate brand style.
        ws.cell(row=r, column=1, value=d).style = "BrandDocsDate"
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=prod)
        ac = ws.cell(row=r, column=4, value=amt)
        ac.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    # A total row with a SUM the generator must preserve (brand total style).
    ws.cell(row=5, column=3, value="Total").style = "BrandDocsTotal"
    tot = ws.cell(row=5, column=4, value="=SUM(D2:D4)")
    tot.style = "BrandDocsTotal"
    tot.number_format = "#,##0"
    ws.freeze_panes = "A2"
    # A date validation on the date column with an input prompt + error alert.
    dv_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="2000-01-01",
        allow_blank=False,
        showInputMessage=True,
        showErrorMessage=True,
        promptTitle="Transaction date",
        prompt="Enter the transaction date (ISO yyyy-mm-dd).",
        errorTitle="Invalid date",
        error="Date must be on or after 2000-01-01.",
    )
    ws.add_data_validation(dv_date)
    dv_date.add("A2:A4")
    # CF families spread to the Data sheet: flag duplicate region values and
    # highlight the 'Widget' product rows (containsText), with brand dxf fills.
    ws.conditional_formatting.add(
        "B2:B4",
        Rule(
            type="duplicateValues",
            dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=BRAND_LIGHT)),
        ),
    )
    ws.conditional_formatting.add(
        "C2:C4",
        Rule(
            type="containsText",
            operator="containsText",
            text="Widget",
            formula=['NOT(ISERROR(SEARCH("Widget",C2)))'],
            dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=BRAND_LIGHT)),
        ),
    )
    for col, w in zip("ABCD", (14, 12, 14, 16)):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Named ranges (workbook scope) - the author's OWN vocabulary, surfaced as
# geometry by the extractor (never matched on as code literals).
# ---------------------------------------------------------------------------
def _add_named_ranges(wb: Workbook) -> None:
    defns = {
        # Single-cell title anchor under a merged header band.
        "report_title": "'Cover'!$A$1",
        "report_subtitle": "'Cover'!$A$2",
        "client_name": "'Cover'!$B$4",
        "period": "'Cover'!$B$5",
        # Multi-cell INPUTS block (sample-data candidate, in the frozen band edge).
        "inputs_block": "'Inputs'!$A$4:$C$7",
        # Multi-cell DATA region = body of the native table.
        "model_body": "'Model'!$A$4:$G$6",
        # A cross-sheet formula OUTPUT cell (single-cell named output slot).
        "headline_kpi": "'Summary'!$B$9",
        "dashboard_kpis": "'Dashboard'!$A$4:$H$5",
        "scenario_block": "'Scenarios'!$A$5:$D$8",
        # The raw demo-data block.
        "data_block": "'Data'!$A$2:$D$4",
    }
    for name in sorted(defns):
        wb.defined_names.add(DefinedName(name, attr_text=defns[name]))
    # A SHEET-SCOPED defined name (local to the Model sheet) pointing at the FY
    # total cell, alongside the workbook-scope names above. localSheetId is the
    # 0-based index of the Model worksheet, so the name resolves only on that tab.
    model_idx = wb.sheetnames.index("Model")
    wb["Model"].defined_names.add(
        DefinedName("q_total", attr_text="Model!$F$7", localSheetId=model_idx)
    )


_LOGO_CACHE: Path | None = None


def _logo_path() -> Path:
    """Materialize the shared BrandDocs wordmark to a temp file openpyxl can
    embed.

    Uses ``branddocs_mark_png`` so the cover logo is the SAME generated
    text-only wordmark used by the DOCX and PPTX examples.
    """
    global _LOGO_CACHE
    if _LOGO_CACHE is None:
        import tempfile

        tmp = Path(tempfile.gettempdir()) / "branddocs_template_wordmark.png"
        tmp.write_bytes(branddocs_mark_png(640, 160))
        _LOGO_CACHE = tmp
    return _LOGO_CACHE


# ---------------------------------------------------------------------------
# Raw-package post-processing (parts openpyxl's object model cannot author):
#   * theme1.xml a:clrScheme + a:fontScheme  -> the workbook-level BrandDocs palette
#   * styles.xml  custom dxf-backed table style 'BrandDocsTableStyle'
#   * the Dashboard sheet's x14 sparklineGroups (in-cell line trends)
# All three are STATIC rewrites of an existing saved part (constants only, no
# wall-clock), performed BEFORE freeze_ooxml pins the timestamps. The steps are
# order-stable, so two builds stay byte-identical.
# ---------------------------------------------------------------------------
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_S_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_REVISION_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# Brand hexes (bare RRGGBB) for the raw-XML parts, sourced from the shared
# _brandlib slot map so the .xlsx theme can never drift from the .docx/.pptx one.
_SLOTS = brand_theme_slots()
_NAVY = _SLOTS["dk1"]  # 16213F
_TEAL = _SLOTS["accent1"]  # 2B7CD3
_AMBER = _SLOTS["accent2"]  # E0742B
_LIGHTHEX = _SLOTS["lt1"]  # EAF1FF

# Full BrandDocs clrScheme: navy text, white/light surfaces, teal primary, amber
# danger, a light tint + two supporting blues for accent3-6, teal/navy hyperlinks.
_CLRSCHEME = {
    "dk1": _NAVY,
    "lt1": "FFFFFF",
    "dk2": _TEAL,
    "lt2": _LIGHTHEX,
    "accent1": _TEAL,
    "accent2": _AMBER,
    "accent3": "DCE7FF",
    "accent4": "5778B0",
    "accent5": "9CC0F0",
    "accent6": "C7912B",
    "hlink": _TEAL,
    "folHlink": _NAVY,
}


def _rewrite_theme(xml: bytes) -> bytes:
    """Rewrite theme1.xml's a:clrScheme to the BrandDocs palette + Arial/Calibri.

    dk1/lt1 are authored as plain srgbClr (no sysClr), so the parsed slot carries
    the brand hex directly. majorFont latin -> Arial, minorFont latin -> Calibri.
    """
    root = etree.fromstring(xml)

    def a(tag: str) -> str:
        return f"{{{_A_NS}}}{tag}"

    scheme = root.find(f".//{a('clrScheme')}")
    if scheme is not None:
        for slot, hexval in _CLRSCHEME.items():
            node = scheme.find(a(slot))
            if node is None:
                continue
            for child in list(node):
                node.remove(child)
            srgb = etree.SubElement(node, a("srgbClr"))
            srgb.set("val", hexval)
    font_scheme = root.find(f".//{a('fontScheme')}")
    if font_scheme is not None:
        major = font_scheme.find(a("majorFont"))
        minor = font_scheme.find(a("minorFont"))
        for fonts, face in ((major, "Arial"), (minor, "Calibri")):
            if fonts is None:
                continue
            latin = fonts.find(a("latin"))
            if latin is None:
                latin = etree.SubElement(fonts, a("latin"))
            latin.set("typeface", face)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


# Custom table-style dxf bands (header navy/white-bold, first-row stripe brand
# light, totals teal/white-bold). Injected into styles.xml as new <dxf> entries
# plus a <tableStyle name="BrandDocsTableStyle"> referencing them.
def _inject_table_style(xml: bytes) -> bytes:
    """Add the BrandDocsTableStyle (header/stripe/total dxf bands) to styles.xml."""
    root = etree.fromstring(xml)

    def s(tag: str) -> str:
        return f"{{{_S_NS}}}{tag}"

    # 1) Append three brand dxfs (header / first-row-stripe / total) and remember
    #    their indices (dxfId is the 0-based position in <dxfs>).
    dxfs = root.find(s("dxfs"))
    if dxfs is None:
        dxfs = etree.Element(s("dxfs"))
        # <dxfs> sits just before <tableStyles> in the schema order; insert near end.
        root.append(dxfs)
    base = len(dxfs)

    def _band_dxf(font_hex: str | None, fill_hex: str, bold: bool) -> None:
        dxf = etree.SubElement(dxfs, s("dxf"))
        if font_hex is not None:
            font = etree.SubElement(dxf, s("font"))
            if bold:
                etree.SubElement(font, s("b"))
            etree.SubElement(font, s("color")).set("rgb", "FF" + font_hex)
        fill = etree.SubElement(dxf, s("fill"))
        pattern = etree.SubElement(fill, s("patternFill"))
        etree.SubElement(pattern, s("fgColor")).set("rgb", "FF" + fill_hex)
        etree.SubElement(pattern, s("bgColor")).set("rgb", "FF" + fill_hex)

    _band_dxf("FFFFFF", _NAVY, True)  # header band
    _band_dxf(None, _LIGHTHEX, False)  # first-row stripe
    _band_dxf("FFFFFF", _TEAL, True)  # totals band
    header_id, stripe_id, total_id = base, base + 1, base + 2
    dxfs.set("count", str(len(dxfs)))

    # 2) Register the named tableStyle referencing those dxfs.
    table_styles = root.find(s("tableStyles"))
    if table_styles is None:
        table_styles = etree.SubElement(root, s("tableStyles"))
    style = etree.SubElement(table_styles, s("tableStyle"))
    style.set("name", "BrandDocsTableStyle")
    style.set("pivot", "0")
    style.set("count", "3")
    for etype, dxf_id in (
        ("headerRow", header_id),
        ("firstRowStripe", stripe_id),
        ("totalRow", total_id),
    ):
        el = etree.SubElement(style, s("tableStyleElement"))
        el.set("type", etype)
        el.set("dxfId", str(dxf_id))
    existing = int(table_styles.get("count", "0") or "0")
    table_styles.set("count", str(existing + 1))
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


# In-cell line sparklines over the Dashboard KPI-trend mini-table (one group per
# KPI row across its Q1..Q4 cells, drawn in column F). openpyxl has no sparkline
# model, so the x14 extension is injected into the sheet part directly.
_SPARKLINES = [
    ("F16", "B16:E16"),
    ("F17", "B17:E17"),
    ("F18", "B18:E18"),
]


def _inject_sparklines(xml: bytes) -> bytes:
    """Append an x14 sparklineGroups extension to the Dashboard worksheet part.

    The injected subtree carries an EXPLICIT ``nsmap`` (``x14`` + ``xm`` prefixes)
    declared once on the ``sparklineGroups`` element, so the serialization is valid
    and byte-stable (no per-element auto-prefixing, no reserved-namespace binding).
    """
    root = etree.fromstring(xml)

    def s(tag: str) -> str:
        return f"{{{_S_NS}}}{tag}"

    def x14(tag: str) -> str:
        return f"{{{_X14_NS}}}{tag}"

    def xm(tag: str) -> str:
        return f"{{{_XM_NS}}}{tag}"

    ext_lst = root.find(s("extLst"))
    if ext_lst is None:
        ext_lst = etree.SubElement(root, s("extLst"))
    ext = etree.SubElement(ext_lst, s("ext"))
    ext.set(f"{{{_REVISION_NS}}}id", "rIdSparkline")
    ext.set("uri", "{05C60535-1F16-4fd2-B633-F4F36F0B64E0}")
    groups = etree.SubElement(
        ext, x14("sparklineGroups"), nsmap={"x14": _X14_NS, "xm": _XM_NS}
    )
    for loc, data_range in _SPARKLINES:
        group = etree.SubElement(groups, x14("sparklineGroup"))
        group.set("displayEmptyCellsAs", "gap")
        group.set("type", "line")
        etree.SubElement(group, x14("colorSeries")).set("rgb", "FF" + _TEAL)
        etree.SubElement(group, x14("colorNegative")).set("rgb", "FF" + _AMBER)
        etree.SubElement(group, x14("colorMarkers")).set("rgb", "FF" + _NAVY)
        sparks = etree.SubElement(group, x14("sparklines"))
        spark = etree.SubElement(sparks, x14("sparkline"))
        etree.SubElement(spark, xm("f")).text = f"Dashboard!{data_range}"
        etree.SubElement(spark, xm("sqref")).text = loc
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _sheet_part_for(out: Path, sheet_name: str) -> str:
    """Resolve a worksheet display name to its ``xl/worksheets/sheetN.xml`` part.

    Reads workbook.xml + its rels so the mapping is robust to sheet reordering
    (never hardcodes sheetN). Returns the part path (without leading slash).
    """
    with zipfile.ZipFile(out) as z:
        wbxml = etree.fromstring(z.read("xl/workbook.xml"))
        rels = etree.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    s_ns = _S_NS
    rid = None
    for sheet in wbxml.find(f"{{{s_ns}}}sheets"):
        if sheet.get("name") == sheet_name:
            rid = sheet.get(f"{{{r_ns}}}id")
            break
    if rid is None:
        raise KeyError(f"sheet {sheet_name!r} not found in workbook.xml")
    target = None
    for rel in rels:
        if rel.get("Id") == rid:
            target = rel.get("Target")
            break
    if target is None:
        raise KeyError(f"relationship {rid!r} not found")
    target = target.lstrip("/")
    if not target.startswith("xl/"):
        target = "xl/" + target
    return target


def _post_process_package(out: Path) -> None:
    """Apply the static raw-XML rewrites in a single deterministic zip rewrite."""
    dashboard_part = _sheet_part_for(out, "Dashboard")
    with open(out, "rb") as fh:
        data = fh.read()
    buf = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(data)) as zin,
        zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout,
    ):
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/theme/theme1.xml":
                payload = _rewrite_theme(payload)
            elif info.filename == "xl/styles.xml":
                payload = _inject_table_style(payload)
            elif info.filename == dashboard_part:
                payload = _inject_sparklines(payload)
            zout.writestr(info, payload)
    with open(out, "wb") as fh:
        fh.write(buf.getvalue())


def build(out: Path = OUT) -> Path:
    wb = Workbook()
    # Drop the default sheet; we author named sheets in a deliberate order.
    wb.remove(wb.active)
    _register_named_styles(wb)
    _build_cover(wb)
    _build_inputs(wb)
    _build_model(wb)
    _build_summary(wb)
    _build_dashboard(wb)
    _build_scenarios(wb)
    _build_data(wb)
    _add_named_ranges(wb)
    # Workbook-level: request a full recalc so authored formulas evaluate on open.
    wb.calculation.fullCalcOnLoad = True
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    # Static raw-XML enrichments openpyxl cannot model (theme palette/fonts, the
    # custom dxf table style, the in-cell sparklines), BEFORE freeze_ooxml.
    _post_process_package(out)
    freeze_ooxml(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"built {path}")
